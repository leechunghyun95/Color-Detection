# 이미지 데이터가 담긴 엑셀파일(img_list.xlsx)
# 해당 파일에서 D4('rep_image_path') 칼럼값만 가져와서 배열에 담는다.
# 배열을 돌며 url + rep_image_path값을 요청하여 로컬에 이미지를 저장한다.

from openpyxl import load_workbook
from urllib import request

url = "https://little-picasso-prod.s3.ap-northeast-2.amazonaws.com"

# img_list.xlsx 파일을 연다
img_list_excelFile = load_workbook('img_list.xlsx', data_only=True)

# img_list에서 시트를 선택하고 artworkListSheet변수에 담는다.
artworkListSheet = img_list_excelFile['아트워크 리스트']

# artworkListSheet의 열의 개수를 가져온다.
row_num = artworkListSheet.max_row

# 반복문을 돌며 D2 ~ D501까지의 rep_image_path를 가져온다.
# url + rep_image_path를 하여 이미지를 요청하고 저장한다.
for i in range(2,row_num + 1):
    rep_image_path_location = 'D' + str(i)
    rep_image_path = artworkListSheet[rep_image_path_location].value
    
    image_path = url + rep_image_path
    image_name ='./img/' + str(i-1) + '.png'

    request.urlretrieve(image_path,image_name)



